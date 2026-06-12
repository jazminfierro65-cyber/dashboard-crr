from pathlib import Path

import openpyxl

source = Path("/Users/jazz/Downloads/plantilla_carga_web_crr.xlsx")
backup = Path("/Users/jazz/Downloads/plantilla_carga_web_crr_backup_antes_arreglo.xlsx")
dashboard_copy = Path("/Users/jazz/Documents/New project/outputs/crr_dashboard/plantilla_carga_web_crr.xlsx")


def clean(value):
    return "" if value is None else str(value).strip()


def norm_code(value):
    text = clean(value).split("-")[0]
    return "".join(ch for ch in text if ch.isdigit())


def time_minutes(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return round((float(value) % 1) * 24 * 60)
    if hasattr(value, "hour") and hasattr(value, "minute"):
        return value.hour * 60 + value.minute
    text = clean(value)
    if ":" not in text:
        return None
    try:
        hour, minute = text.split(":")[:2]
        return int(hour) * 60 + int(minute)
    except ValueError:
        return None


def minutes_between(start, end):
    a = time_minutes(start)
    b = time_minutes(end)
    if a is None or b is None:
        return None
    diff = b - a
    return diff if diff >= 0 else diff + 1440


def entry_class(entry):
    minutes = time_minutes(entry)
    if minutes is None:
        return None
    return "Ingreso adecuado" if minutes <= 8 * 60 + 15 else "Atraso"


def category_maps():
    categories_by_episode = {}
    if backup.exists():
        old_wb = openpyxl.load_workbook(backup, data_only=True)
        old_ws = old_wb["Carga Web Cirugias"]
        for row in range(4, old_ws.max_row + 1):
            episode = clean(old_ws.cell(row, 2).value)
            category = clean(old_ws.cell(row, 10).value)
            if episode and category:
                categories_by_episode[episode] = category
    code_map = {}
    wb_for_catalog = openpyxl.load_workbook(source, data_only=True)
    if "Catalogo MINSAL" in wb_for_catalog.sheetnames:
        cat_ws = wb_for_catalog["Catalogo MINSAL"]
        for row in range(4, cat_ws.max_row + 1):
            code = norm_code(cat_ws.cell(row, 1).value)
            category = clean(cat_ws.cell(row, 5).value)
            if code and category:
                code_map[code] = category
    return categories_by_episode, code_map


episode_category, code_category = category_maps()

wb = openpyxl.load_workbook(source)
ws = wb["Carga Web Cirugias"]

for row in range(4, 501):
    has_data = any(clean(ws.cell(row, col).value) for col in range(1, 29))
    if not has_data:
        continue

    episode = clean(ws.cell(row, 2).value)
    if episode_category.get(episode):
        ws.cell(row, 10).value = episode_category[episode]
    else:
        category = code_category.get(norm_code(ws.cell(row, 8).value)) or code_category.get(norm_code(ws.cell(row, 9).value))
        if category:
            ws.cell(row, 10).value = category

    ws.cell(row, 18).value = entry_class(ws.cell(row, 17).value)
    ws.cell(row, 22).value = minutes_between(ws.cell(row, 17).value, ws.cell(row, 19).value)
    ws.cell(row, 23).value = minutes_between(ws.cell(row, 19).value, ws.cell(row, 20).value)
    ws.cell(row, 24).value = None
    ws.cell(row, 25).value = None
    ws.cell(row, 26).value = None

    ws.cell(row, 17).number_format = "hh:mm"
    for col in (19, 20, 21):
        ws.cell(row, col).number_format = "hh:mm"
    for col in (22, 23, 24, 25):
        ws.cell(row, col).number_format = "0"

dashboard_copy.parent.mkdir(parents=True, exist_ok=True)
wb.save(source)
wb.save(dashboard_copy)
print({"fixed": str(source), "dashboard_copy": str(dashboard_copy)})
