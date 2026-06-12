from pathlib import Path

import openpyxl

template_path = Path("/Users/jazz/Documents/New project/plantilla_carga_web_crr.xltx")

values_wb = openpyxl.load_workbook(template_path, data_only=True)
values_ws = values_wb["Carga Web Cirugias"]
programmed_values = [values_ws[f"X{row}"].value for row in range(4, 501)]

wb = openpyxl.load_workbook(template_path)
ws = wb["Carga Web Cirugias"]

for offset, row in enumerate(range(4, 501)):
    ws[f"V{row}"] = (
        f'=IFERROR(IF(OR(Q{row}="",S{row}=""),"",'
        f'IF(IF(ISNUMBER(S{row}),MOD(S{row},1),TIMEVALUE(S{row}))>=IF(ISNUMBER(Q{row}),MOD(Q{row},1),TIMEVALUE(Q{row})),'
        f'(IF(ISNUMBER(S{row}),MOD(S{row},1),TIMEVALUE(S{row}))-IF(ISNUMBER(Q{row}),MOD(Q{row},1),TIMEVALUE(Q{row})))*1440,'
        f'(IF(ISNUMBER(S{row}),MOD(S{row},1),TIMEVALUE(S{row}))+1-IF(ISNUMBER(Q{row}),MOD(Q{row},1),TIMEVALUE(Q{row})))*1440)),"")'
    )
    ws[f"W{row}"] = (
        f'=IFERROR(IF(OR(S{row}="",T{row}=""),"",'
        f'IF(IF(ISNUMBER(T{row}),MOD(T{row},1),TIMEVALUE(T{row}))>=IF(ISNUMBER(S{row}),MOD(S{row},1),TIMEVALUE(S{row})),'
        f'(IF(ISNUMBER(T{row}),MOD(T{row},1),TIMEVALUE(T{row}))-IF(ISNUMBER(S{row}),MOD(S{row},1),TIMEVALUE(S{row})))*1440,'
        f'(IF(ISNUMBER(T{row}),MOD(T{row},1),TIMEVALUE(T{row}))+1-IF(ISNUMBER(S{row}),MOD(S{row},1),TIMEVALUE(S{row})))*1440)),"")'
    )
    ws[f"X{row}"] = programmed_values[offset]
    ws[f"Y{row}"] = f'=IFERROR(IF(OR(W{row}="",X{row}=""),"",W{row}-X{row}),"")'
    ws[f"Z{row}"] = f'=IFERROR(IF(Y{row}="","",IF(Y{row}>15,"Subestimacion",IF(Y{row}<-15,"Sobrestimacion","Asertivo"))),"")'

for row in range(4, 501):
    ws[f"Q{row}"].number_format = "hh:mm"
    ws[f"S{row}"].number_format = "hh:mm"
    ws[f"T{row}"].number_format = "hh:mm"
    ws[f"U{row}"].number_format = "hh:mm"
    for col in ("V", "W", "X", "Y"):
        ws[f"{col}{row}"].number_format = "0"

wb.template = True
wb.save(template_path)
print(template_path)
