from pathlib import Path

import openpyxl

source = Path("/Users/jazz/Downloads/plantilla_carga_web_crr.xlsx")
output_for_dashboard = Path("/Users/jazz/Documents/New project/outputs/crr_dashboard/plantilla_carga_web_crr.xlsx")


def clean(value):
    return "" if value is None else str(value).strip()


def norm_code(value):
    text = clean(value)
    if not text:
        return ""
    text = text.split("-")[0]
    return "".join(ch for ch in text if ch.isdigit())


def non_empty_count(ws, row):
    return sum(1 for col in range(1, 29) if clean(ws.cell(row, col).value))


def delay_formula(row):
    return (
        f'=IFERROR(IF(OR(Q{row}="",S{row}=""),"",'
        f'IF(IF(ISNUMBER(S{row}),MOD(S{row},1),TIMEVALUE(S{row}))>=IF(ISNUMBER(Q{row}),MOD(Q{row},1),TIMEVALUE(Q{row})),'
        f'(IF(ISNUMBER(S{row}),MOD(S{row},1),TIMEVALUE(S{row}))-IF(ISNUMBER(Q{row}),MOD(Q{row},1),TIMEVALUE(Q{row})))*1440,'
        f'(IF(ISNUMBER(S{row}),MOD(S{row},1),TIMEVALUE(S{row}))+1-IF(ISNUMBER(Q{row}),MOD(Q{row},1),TIMEVALUE(Q{row})))*1440)),"")'
    )


def duration_formula(row):
    return (
        f'=IFERROR(IF(OR(S{row}="",T{row}=""),"",'
        f'IF(IF(ISNUMBER(T{row}),MOD(T{row},1),TIMEVALUE(T{row}))>=IF(ISNUMBER(S{row}),MOD(S{row},1),TIMEVALUE(S{row})),'
        f'(IF(ISNUMBER(T{row}),MOD(T{row},1),TIMEVALUE(T{row}))-IF(ISNUMBER(S{row}),MOD(S{row},1),TIMEVALUE(S{row})))*1440,'
        f'(IF(ISNUMBER(T{row}),MOD(T{row},1),TIMEVALUE(T{row}))+1-IF(ISNUMBER(S{row}),MOD(S{row},1),TIMEVALUE(S{row})))*1440)),"")'
    )


def diff_formula(row):
    return f'=IFERROR(IF(OR(W{row}="",X{row}=""),"",W{row}-X{row}),"")'


def accuracy_formula(row):
    return f'=IFERROR(IF(Y{row}="","",IF(Y{row}>15,"Subestimacion",IF(Y{row}<-15,"Sobrestimacion","Asertivo"))),"")'


wb = openpyxl.load_workbook(source)
ws = wb["Carga Web Cirugias"]

category_by_code = {}
if "Catalogo MINSAL" in wb.sheetnames:
    cat_ws = wb["Catalogo MINSAL"]
    for row in range(4, cat_ws.max_row + 1):
        code = norm_code(cat_ws.cell(row, 1).value)
        category = clean(cat_ws.cell(row, 5).value)
        if code and category:
            category_by_code[code] = category

episodes = {}
for row in range(4, ws.max_row + 1):
    episode = clean(ws.cell(row, 2).value)
    has_data = any(clean(ws.cell(row, col).value) for col in range(1, 29))
    if episode and has_data:
        episodes.setdefault(episode, []).append(row)

rows_to_clear = set()
for rows in episodes.values():
    if len(rows) <= 1:
        continue
    keep = max(rows, key=lambda r: (non_empty_count(ws, r), r))
    rows_to_clear.update(r for r in rows if r != keep)

for row in rows_to_clear:
    for col in range(1, 29):
        ws.cell(row, col).value = None

for row in range(4, 501):
    has_data = any(clean(ws.cell(row, col).value) for col in range(1, 29))
    if has_data:
        if not clean(ws.cell(row, 10).value):
            category = category_by_code.get(norm_code(ws.cell(row, 8).value)) or category_by_code.get(norm_code(ws.cell(row, 9).value))
            if category:
                ws.cell(row, 10).value = category
        ws.cell(row, 24).value = None
    ws.cell(row, 22).value = delay_formula(row)
    ws.cell(row, 23).value = duration_formula(row)
    ws.cell(row, 25).value = diff_formula(row)
    ws.cell(row, 26).value = accuracy_formula(row)

    ws.cell(row, 17).number_format = "hh:mm"
    for col in (19, 20, 21):
        ws.cell(row, col).number_format = "hh:mm"
    for col in (22, 23, 24, 25):
        ws.cell(row, col).number_format = "0"

output_for_dashboard.parent.mkdir(parents=True, exist_ok=True)
wb.save(source)
wb.save(output_for_dashboard)
print(
    {
        "fixed": str(source),
        "dashboard_copy": str(output_for_dashboard),
        "duplicate_rows_cleared": len(rows_to_clear),
    }
)
